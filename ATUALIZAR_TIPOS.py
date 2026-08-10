#!/usr/bin/env python3
"""
Atualiza a coluna Tipo do ClubeDoCinema com os dados da auditoria TMDb
"""

import openpyxl

# Arquivos
AUDITORIA = "AUDITORIA_TMDB_RESULTADO_FINAL.xlsx"
PLANILHA_ORIGINAL = "ClubeDoCinema - Copia (2).xlsx"

print("=" * 70)
print("🔄 ATUALIZANDO TIPOS NA PLANILHA")
print("=" * 70)

# Lê auditoria
print(f"\n📖 Lendo auditoria: {AUDITORIA}")
wb_auditoria = openpyxl.load_workbook(AUDITORIA)
ws_auditoria = wb_auditoria.active

auditorias = {}
for row in ws_auditoria.iter_rows(min_row=2, values_only=True):
    numero = row[0]  # Coluna A: Número
    titulo = row[1]  # Coluna B: Título
    tipo_atual = row[2]  # Coluna C: Tipo Atual
    episodios = row[3]  # Coluna D: Episódios (TMDb)
    tipo_correto = row[4]  # Coluna E: Tipo Correto
    status = row[5]  # Coluna F: Status

    if numero and titulo:
        auditorias[str(numero).strip()] = {
            'titulo': titulo,
            'tipo_atual': tipo_atual,
            'tipo_correto': tipo_correto,
            'status': status
        }

print(f"✅ {len(auditorias)} títulos carregados da auditoria")

# Lê planilha original
print(f"\n📖 Lendo planilha: {PLANILHA_ORIGINAL}")
wb_original = openpyxl.load_workbook(PLANILHA_ORIGINAL)
ws_original = wb_original.active

# Encontra coluna Tipo
col_numero = None
col_tipo = None
for col in range(1, 10):
    cell_value = ws_original.cell(1, col).value
    if cell_value and "número" in str(cell_value).lower():
        col_numero = col
    if cell_value and "tipo" in str(cell_value).lower():
        col_tipo = col

if not col_numero or not col_tipo:
    print("❌ Não consegui encontrar colunas 'Número' ou 'Tipo'")
    exit(1)

print(f"✅ Colunas encontradas: Número={col_numero}, Tipo={col_tipo}")

# Atualiza tipos - APENAS S para MS (séries com ≤10 episódios)
atualizados = 0
print(f"\n🔄 Atualizando S → MS (séries com ≤10 episódios)...\n")

for row in range(2, ws_original.max_row + 1):
    numero = ws_original.cell(row, col_numero).value
    tipo_atual = ws_original.cell(row, col_tipo).value

    if numero:
        numero_str = str(numero).strip()

        if numero_str in auditorias:
            audit = auditorias[numero_str]
            tipo_correto = audit['tipo_correto']
            episodios = audit['status']

            # APENAS corrige S → MS (tipo_atual = S, tipo_correto = MS)
            if tipo_atual == 'S' and tipo_correto == 'MS':
                print(f"  {numero_str}: {audit['titulo']}")
                print(f"    S → MS ({audit['tipo_atual']} episódios)")

                ws_original.cell(row, col_tipo).value = 'MS'
                atualizados += 1

print(f"\n{'=' * 70}")
print(f"✅ {atualizados} títulos foram corrigidos")
print(f"{'=' * 70}")

# Salva
wb_original.save(PLANILHA_ORIGINAL)
print(f"\n💾 Planilha atualizada: {PLANILHA_ORIGINAL}")
