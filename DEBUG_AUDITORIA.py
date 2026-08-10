#!/usr/bin/env python3
"""
Debug: verifica o match entre auditoria e planilha
"""

import openpyxl

AUDITORIA = "AUDITORIA_TMDB_RESULTADO_FINAL.xlsx"
PLANILHA = "ClubeDoCinema - Copia (2).xlsx"

print("=" * 70)
print("🔍 DEBUG: VERIFICANDO MATCH")
print("=" * 70)

# Lê auditoria
print(f"\n📖 Lendo: {AUDITORIA}")
wb_a = openpyxl.load_workbook(AUDITORIA)
ws_a = wb_a.active

numeros_auditoria = set()
s_para_ms = 0

for row in ws_a.iter_rows(min_row=2, values_only=True):
    numero = row[0]
    tipo_atual = row[2]
    tipo_correto = row[3]

    if numero:
        numeros_auditoria.add(str(numero).strip())
        # Conta S→MS
        if tipo_atual == 'S' and tipo_correto == 'MS':
            s_para_ms += 1

print(f"✅ {len(numeros_auditoria)} números únicos na auditoria")
print(f"   → {s_para_ms} casos de S→MS encontrados")

# Lê planilha
print(f"\n📖 Lendo: {PLANILHA}")
wb_p = openpyxl.load_workbook(PLANILHA)
ws_p = wb_p.active

numeros_planilha = set()
tipos_s = []
tipos_ms = []

for row in range(2, ws_p.max_row + 1):
    numero = ws_p.cell(row, 1).value
    tipo = ws_p.cell(row, 3).value

    if numero:
        numero_str = str(numero).strip()
        numeros_planilha.add(numero_str)

        if tipo == 'S':
            tipos_s.append((numero_str, ws_p.cell(row, 2).value))
        elif tipo == 'MS':
            tipos_ms.append((numero_str, ws_p.cell(row, 2).value))

print(f"✅ {len(numeros_planilha)} números únicos na planilha")
print(f"   → {len(tipos_s)} com tipo S")
print(f"   → {len(tipos_ms)} com tipo MS")

# Compara
match = numeros_auditoria & numeros_planilha
nao_match = numeros_auditoria - numeros_planilha

print(f"\n🔗 Match: {len(match)} números encontram em ambos")
print(f"❌ Não encontram na planilha: {len(nao_match)}")

if nao_match and len(nao_match) <= 20:
    print(f"\nNúmeros na auditoria mas NÃO na planilha:")
    for n in sorted(list(nao_match)):
        print(f"  - {n}")

# Mostra S→MS da auditoria
print(f"\n📋 S→MS na auditoria:")
count = 0
for row in ws_a.iter_rows(min_row=2, values_only=True):
    numero = row[0]
    titulo = row[1]
    tipo_atual = row[2]
    tipo_correto = row[3]

    if tipo_atual == 'S' and tipo_correto == 'MS' and count < 10:
        print(f"  {numero}: {titulo} (S→MS)")
        count += 1
if s_para_ms > 10:
    print(f"  ... e mais {s_para_ms - 10}")
