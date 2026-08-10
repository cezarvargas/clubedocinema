#!/usr/bin/env python3
"""
AUDITORIA DE SÉRIES E MINISSÉRIES - TMDB
==========================================
Verifica todos os títulos S/MS na planilha ClubeDoCinema.xlsx
Busca episódios reais no TMDb e identifica classificações incorretas.

Classificação Correta:
  - MS (Minissérie): <= 10 episódios
  - S (Série): >= 11 episódios
"""

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import time
import sys

# ===== CONFIGURAÇÃO =====
TMDB_API_KEY = "3cc4fa0abe696f59de02fd922e2a24f0"
PLANILHA_ENTRADA = "ClubeDoCinema.xlsx"
PLANILHA_SAIDA = "AUDITORIA_TMDB_RESULTADO_FINAL.xlsx"

# Cores para o relatório
COR_CORRETO = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde
COR_INCORRETO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Vermelho
COR_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Azul

FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
FONT_NORMAL = Font(size=10)
FONT_ERRO = Font(size=10, color="CC0000")

# ===== FUNÇÕES TMDb =====

def buscar_no_tmdb(titulo, tentativa=1):
    """Busca uma série no TMDb e retorna ID + número de episódios."""
    if tentativa > 3:
        return None

    url = "https://api.themoviedb.org/3/search/tv"
    params = {
        "api_key": TMDB_API_KEY,
        "query": titulo,
        "language": "pt-BR"
    }

    try:
        resposta = requests.get(url, params=params, timeout=10)
        resposta.raise_for_status()

        resultados = resposta.json().get("results", [])

        if not resultados:
            # Tenta novamente sem acentos
            if tentativa == 1:
                titulo_limpo = titulo.replace("ã", "a").replace("é", "e").replace("ó", "o").replace("ç", "c")
                if titulo_limpo != titulo:
                    return buscar_no_tmdb(titulo_limpo, tentativa=2)
            return None

        serie = resultados[0]
        id_tmdb = serie.get("id")

        # Busca detalhes (número de episódios)
        url_detalhes = f"https://api.themoviedb.org/3/tv/{id_tmdb}"
        params_detalhe = {
            "api_key": TMDB_API_KEY,
            "language": "pt-BR"
        }

        resp_detalhe = requests.get(url_detalhes, params=params_detalhe, timeout=10)
        resp_detalhe.raise_for_status()

        dados = resp_detalhe.json()
        episodios = dados.get("number_of_episodes", 0)
        nome_tmdb = dados.get("name", "")

        return {
            "id": id_tmdb,
            "episodios": episodios,
            "nome_tmdb": nome_tmdb
        }

    except Exception as e:
        print(f"  ❌ Erro ao buscar '{titulo}': {str(e)}")
        time.sleep(1)
        return None


def classificar_corretamente(episodios):
    """Retorna a classificação correta baseada no número de episódios."""
    if episodios <= 10:
        return "MS"
    else:
        return "S"


# ===== LEITURA DA PLANILHA =====

def ler_planilha():
    """Lê a planilha e extrai títulos S/MS."""
    print(f"\n📖 Lendo planilha: {PLANILHA_ENTRADA}")

    wb = openpyxl.load_workbook(PLANILHA_ENTRADA)
    ws = wb.active

    titulos = []

    # Estrutura fixa da planilha
    col_titulo = 2   # Coluna B
    col_tipo = 3     # Coluna C
    col_ano = 4      # Coluna D

    print(f"✅ Usando colunas: Título (col {col_titulo}), Tipo (col {col_tipo}), Ano (col {col_ano})")

    # Extrai títulos S/MS
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if col_titulo <= len(row) and col_tipo <= len(row):
            numero = row[0]  # Coluna A = Número
            titulo = row[col_titulo - 1]
            tipo = row[col_tipo - 1]
            ano = row[col_ano - 1] if col_ano and col_ano <= len(row) else None

            if titulo and tipo and str(tipo).strip().upper() in ["S", "MS"]:
                titulos.append({
                    "numero": numero if numero else "?",
                    "titulo": str(titulo).strip(),
                    "tipo": str(tipo).strip().upper(),
                    "ano": int(ano) if ano and str(ano).isdigit() else None,
                    "linha": row_num
                })

    print(f"✅ Encontrados {len(titulos)} títulos S/MS")
    return titulos


# ===== AUDITORIA =====

def auditar():
    """Executa a auditoria completa."""

    print("=" * 70)
    print("🎬 AUDITORIA DE CLASSIFICAÇÃO S/MS - TMDB")
    print("=" * 70)

    # Lê planilha
    titulos = ler_planilha()
    if not titulos:
        print("❌ Nenhum título encontrado")
        sys.exit(1)

    # Processa cada título
    resultados = []
    encontrados = 0
    nao_encontrados = 0
    incorretos = 0

    print(f"\n🔍 Buscando {len(titulos)} títulos no TMDb...\n")

    for i, item in enumerate(titulos, 1):
        print(f"[{i}/{len(titulos)}] {item['titulo']} ({item['tipo']})...", end=" ", flush=True)

        # Busca no TMDb
        tmdb_resultado = buscar_no_tmdb(item["titulo"])

        if tmdb_resultado:
            episodios = tmdb_resultado["episodios"]
            classificacao_correta = classificar_corretamente(episodios)

            is_correto = (item["tipo"] == classificacao_correta)

            if not is_correto:
                incorretos += 1
                print(f"❌ {episodios} ep (deveria ser {classificacao_correta})")
            else:
                print(f"✅ {episodios} ep")

            encontrados += 1

            resultados.append({
                "numero": item["numero"],
                "titulo": item["titulo"],
                "tipo_atual": item["tipo"],
                "episodios_tmdb": episodios,
                "tipo_correto": classificacao_correta,
                "correto": is_correto,
                "nome_tmdb": tmdb_resultado["nome_tmdb"],
                "status": "✅ Encontrado" if is_correto else "❌ ERRO"
            })
        else:
            nao_encontrados += 1
            print("⚠️  Não encontrado no TMDb")

            resultados.append({
                "numero": item["numero"],
                "titulo": item["titulo"],
                "tipo_atual": item["tipo"],
                "episodios_tmdb": "N/A",
                "tipo_correto": "?",
                "correto": None,
                "nome_tmdb": "-",
                "status": "⚠️  Não encontrado"
            })

        # Rate limiting
        time.sleep(0.5)

    # ===== GERA RELATÓRIO =====

    print("\n" + "=" * 70)
    print("📊 GERANDO RELATÓRIO...")
    print("=" * 70)

    wb_saida = openpyxl.Workbook()
    ws_saida = wb_saida.active
    ws_saida.title = "Auditoria"

    # Header
    headers = ["Número", "Título", "Tipo Atual", "Episódios (TMDb)", "Tipo Correto", "Status", "Nome TMDb"]
    for col, header in enumerate(headers, 1):
        cell = ws_saida.cell(1, col)
        cell.value = header
        cell.font = FONT_HEADER
        cell.fill = COR_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Dados
    for row_num, resultado in enumerate(resultados, 2):
        ws_saida.cell(row_num, 1).value = resultado["numero"]
        ws_saida.cell(row_num, 2).value = resultado["titulo"]
        ws_saida.cell(row_num, 3).value = resultado["tipo_atual"]
        ws_saida.cell(row_num, 4).value = resultado["episodios_tmdb"]
        ws_saida.cell(row_num, 5).value = resultado["tipo_correto"]
        ws_saida.cell(row_num, 6).value = resultado["status"]
        ws_saida.cell(row_num, 7).value = resultado["nome_tmdb"]

        # Colorir baseado em correto/incorreto
        for col in range(1, 8):
            cell = ws_saida.cell(row_num, col)
            cell.font = FONT_NORMAL
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            if resultado["correto"] == True:
                cell.fill = COR_CORRETO
            elif resultado["correto"] == False:
                cell.fill = COR_INCORRETO

    # Ajusta larguras
    ws_saida.column_dimensions["A"].width = 10  # Número
    ws_saida.column_dimensions["B"].width = 35  # Título
    ws_saida.column_dimensions["C"].width = 12  # Tipo Atual
    ws_saida.column_dimensions["D"].width = 18  # Episódios
    ws_saida.column_dimensions["E"].width = 13  # Tipo Correto
    ws_saida.column_dimensions["F"].width = 18  # Status
    ws_saida.column_dimensions["G"].width = 30  # Nome TMDb

    # Salva
    wb_saida.save(PLANILHA_SAIDA)
    print(f"✅ Relatório salvo: {PLANILHA_SAIDA}")

    # Abre automaticamente
    import os
    import subprocess
    try:
        if sys.platform == "win32":
            os.startfile(PLANILHA_SAIDA)
        else:
            subprocess.run(["open", PLANILHA_SAIDA])
        print(f"📂 Abrindo arquivo...")
    except:
        print(f"⚠️  Não consegui abrir automaticamente. Abra manualmente: {PLANILHA_SAIDA}")

    # ===== RESUMO =====
    print("\n" + "=" * 70)
    print("📈 RESUMO")
    print("=" * 70)
    print(f"Total de títulos:        {len(titulos)}")
    print(f"Encontrados no TMDb:     {encontrados} ({100*encontrados//len(titulos)}%)")
    print(f"Não encontrados:         {nao_encontrados}")
    print(f"Incorretamente classificados: {incorretos} ({100*incorretos//encontrados if encontrados > 0 else 0}%)")
    print("=" * 70)


if __name__ == "__main__":
    try:
        auditar()
    except KeyboardInterrupt:
        print("\n\n⚠️  Interrompido pelo usuário")
    except Exception as e:
        print(f"\n\n❌ ERRO: {e}")
        import traceback
        traceback.print_exc()
